from tkinter.constants import DISABLED
import tkinter.filedialog as filedialog
import tkinter as tk
from openpyxl.reader.excel import load_workbook
import pandas as pd
from tkinter import messagebox
import os

master = tk.Tk()

def choose_output_directory():
    path = filedialog.askdirectory()
    output_entry.delete(1, tk.END)  # Remove current text in entry
    output_entry.insert(0, path)  # Insert the 'path'

def merge_chosen_files():

    result_filename = filename_entry.get()
    output_directory = output_entry.get()
    
    if output_directory == "":
        messagebox.showerror("Title", "Please choose an output directory for your file")
        return "error"
    if result_filename == "":
        messagebox.showerror("Error", "Please choose a name for your merged file")
        return "error"

    try:
        chosen_files = []
        input_path = filedialog.askopenfilename(multiple=True, filetypes=[("Excel files","*.xlsx")])
        for i in range(len(input_path)):
            chosen_files.append(input_path[i])

        pandas_df_list = []
        for file in chosen_files:
            wb = load_workbook(filename = file)
            ws = wb.active
            ws.insert_rows(1)
            wb.save(file)
            pandas_df_list.append(pd.read_excel(file))
            ws.delete_rows(1)
            wb.save(file)
        
        excel_merged = pd.concat(pandas_df_list, ignore_index=True)
        result_file_directory = f"{output_directory}/{result_filename}.xlsx"
        excel_merged.to_excel(result_file_directory)
        
        wb = load_workbook(filename= result_file_directory)
        ws = wb.active
        ws.delete_rows(1)
        ws.delete_cols(1)
        wb.save(result_file_directory)
        messagebox.showinfo("Done", "Your file is ready.")

        os.chdir(output_directory)
        os.system(f'start excel.exe {result_filename}.xlsx')
    
    except:
        messagebox.showerror("Error", "Something went wrong, please check if you chose the correct files.")

master.title("XLMerger")
top_frame = tk.Frame(master)
bottom_frame = tk.Frame(master)

line = tk.Frame(master, height=1, width=400, bg="grey80", relief='groove')
output_path = tk.Label(bottom_frame, text="Choose output directory:")
output_entry = tk.Entry(bottom_frame, text="", width=40)
browse2 = tk.Button(bottom_frame, text="Browse", command=choose_output_directory)

filename_label = tk.Label(bottom_frame, text="Choose merged filename:")
filename_entry = tk.Entry(bottom_frame, text="", width=40)

begin_button = tk.Button(bottom_frame, text='Choose Excel files to merge', command=merge_chosen_files)

top_frame.pack(side=tk.TOP)
bottom_frame.pack(side=tk.BOTTOM)

output_path.pack(pady=5)
output_entry.pack(pady=5)
browse2.pack(pady=5)

line.pack(pady=10)

filename_label.pack(pady=5)
filename_entry.pack(pady=5)
browse2.pack(pady=5)

begin_button.pack(pady=20, fill=tk.X)

master.mainloop()